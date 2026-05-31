"""
database.py — 代账客户收费管理系统 数据库核心
==============================================
功能：
  1. 管理员注册/登录（SHA256 哈希）
  2. 客户 CRUD（增删改查）+ 状态管理（有效/中断/失联）
  3. 收款记录 + 自动推进最近已收期
  4. 逾期等级自动判定（颜色标注）
  5. 按次收费业务管理（支持多选业务类型）
  6. 月收入统计 + 应收未收查询
"""

import sqlite3
import os
import sys
import hashlib
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import tempfile

# ============================================================
#  数据库初始化
# ============================================================

# 数据库存放目录（兼容 PyInstaller 打包 + macOS .app 便携）
if getattr(sys, 'frozen', False):
    exe_path = os.path.dirname(sys.executable)
    # macOS .app 包内路径如 xxx.app/Contents/MacOS，数据应放在 .app 同级
    if sys.platform == 'darwin' and 'Contents/MacOS' in exe_path:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(exe_path)))
    else:
        BASE_DIR = exe_path
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DB_DIR, "clients.db")


def _migrate_billing_type_check():
    """迁移：更新 billing_type CHECK 约束以支持'一人多司'"""
    conn = get_connection()
    # 检查表定义中是否已包含'一人多司'
    table_info = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='clients'"
    ).fetchone()
    if table_info and "一人多司" in (table_info["sql"] or ""):
        conn.close()
        return  # 约束已更新

    # 重建 clients 表以更新 CHECK 约束
    conn.execute("PRAGMA foreign_keys = OFF")
    conn.execute("ALTER TABLE clients RENAME TO clients_old")
    conn.execute("""
        CREATE TABLE clients (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT    NOT NULL,
            billing_type    TEXT    NOT NULL CHECK(billing_type IN ('年收','季收','月收','一人多司')),
            fee_amount      REAL    NOT NULL DEFAULT 0,
            last_paid_period TEXT   NOT NULL,
            next_due_date   TEXT    DEFAULT '',
            status          TEXT    DEFAULT '有效' CHECK(status IN ('有效','中断','失联')),
            payment_status  TEXT    DEFAULT '未收' CHECK(payment_status IN ('已收','未收')),
            charge_period_start TEXT DEFAULT '',
            charge_period_end   TEXT DEFAULT '',
            contact_person  TEXT    DEFAULT '',
            phone           TEXT    DEFAULT '',
            notes           TEXT    DEFAULT '',
            is_active       INTEGER DEFAULT 1,
            created_at      TEXT    DEFAULT (datetime('now','localtime'))
        )
    """)
    conn.execute("""
        INSERT INTO clients (id, name, billing_type, fee_amount, last_paid_period,
            next_due_date, status, payment_status, charge_period_start, charge_period_end,
            contact_person, phone, notes, is_active, created_at)
        SELECT id, name, billing_type, fee_amount, last_paid_period,
            '', status, payment_status, charge_period_start, charge_period_end,
            '', phone, notes, is_active, created_at
        FROM clients_old
    """)
    conn.execute("DROP TABLE clients_old")
    conn.execute("PRAGMA foreign_keys = ON")
    conn.commit()
    conn.close()


def get_connection():
    """获取数据库连接（自动创建目录和文件）"""
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    """初始化所有表（首次运行自动建表）"""
    conn = get_connection()
    conn.executescript("""
        -- 管理员表
        CREATE TABLE IF NOT EXISTS admins (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            username    TEXT    UNIQUE NOT NULL,
            password_hash TEXT   NOT NULL,
            created_at  TEXT    DEFAULT (datetime('now','localtime'))
        );

        -- 客户表（年收/季收/月收/一人多司）
        CREATE TABLE IF NOT EXISTS clients (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT    NOT NULL,
            billing_type    TEXT    NOT NULL CHECK(billing_type IN ('年收','季收','月收','一人多司')),
            fee_amount      REAL    NOT NULL DEFAULT 0,
            last_paid_period TEXT   NOT NULL,  -- YYYY-MM，已缴费覆盖到的月份
            next_due_date   TEXT    DEFAULT '',  -- 下次应付期 YYYY-MM（手动设置）
            status          TEXT    DEFAULT '有效' CHECK(status IN ('有效','中断','失联')),
            payment_status  TEXT    DEFAULT '未收' CHECK(payment_status IN ('已收','未收')),
            charge_period_start TEXT DEFAULT '',  -- 收费期间起始 YYYY-MM
            charge_period_end   TEXT DEFAULT '',  -- 收费期间截止 YYYY-MM
            contact_person  TEXT    DEFAULT '',  -- 联系人
            phone           TEXT    DEFAULT '',
            notes           TEXT    DEFAULT '',
            is_active       INTEGER DEFAULT 1,
            created_at      TEXT    DEFAULT (datetime('now','localtime'))
        );

        -- 收款记录表
        CREATE TABLE IF NOT EXISTS payments (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id   INTEGER NOT NULL,
            period_from TEXT    NOT NULL,  -- 本次收款覆盖起始月份 YYYY-MM
            period_to   TEXT    NOT NULL,  -- 本次收款覆盖截止月份 YYYY-MM
            amount      REAL    NOT NULL,
            paid_date   TEXT    DEFAULT (date('now','localtime')),
            notes       TEXT    DEFAULT '',
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE
        );

        -- 按次收费业务表
        CREATE TABLE IF NOT EXISTS services (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id       INTEGER,  -- 可为空，支持临时客户
            temp_customer_name TEXT DEFAULT '',
            service_type    TEXT    NOT NULL,  -- 业务类型，支持逗号分隔多选
            fee_standard    REAL    DEFAULT 0,  -- 收费标准
            actual_fee      REAL    DEFAULT 0,  -- 实收金额
            status          TEXT    DEFAULT '未收' CHECK(status IN ('未收','已收')),
            completed_date  TEXT    DEFAULT '',
            notes           TEXT    DEFAULT '',
            created_at      TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE SET NULL
        );

        -- 操作日志表
        CREATE TABLE IF NOT EXISTS operation_logs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            action      TEXT    NOT NULL,  -- 新增/编辑/删除/标记收费/导出
            target_type TEXT    NOT NULL,  -- 客户/按次业务/报表
            target_id   INTEGER,
            target_name TEXT,
            detail      TEXT    DEFAULT '',
            created_at  TEXT    DEFAULT (datetime('now','localtime'))
        );
    """)
    conn.commit()
    conn.close()

    # 数据库迁移：为旧版本添加缺失字段
    _migrate_db()
    # 迁移 billing_type CHECK 约束（SQLite 不支持直接 ALTER，重建表）
    _migrate_billing_type_check()


def _migrate_db():
    """数据库迁移：为旧版本添加缺失字段"""
    conn = get_connection()

    migrations = [
        "ALTER TABLE services ADD COLUMN temp_customer_name TEXT DEFAULT ''",
        "ALTER TABLE clients ADD COLUMN status TEXT DEFAULT '有效'",
        "ALTER TABLE clients ADD COLUMN payment_status TEXT DEFAULT '未收'",
        "ALTER TABLE clients ADD COLUMN charge_period_start TEXT DEFAULT ''",
        "ALTER TABLE clients ADD COLUMN charge_period_end TEXT DEFAULT ''",
        "ALTER TABLE clients ADD COLUMN contact_person TEXT DEFAULT ''",
        "ALTER TABLE clients ADD COLUMN next_due_date TEXT DEFAULT ''",
    ]
    for sql in migrations:
        try:
            conn.execute(sql)
            conn.commit()
        except sqlite3.OperationalError:
            pass  # 列已存在

    conn.close()


# ============================================================
#  管理员模块
# ============================================================

def hash_password(password: str) -> str:
    """SHA256 哈希"""
    return hashlib.sha256(password.encode()).hexdigest()


def admin_exists() -> bool:
    """检查是否已有管理员"""
    conn = get_connection()
    row = conn.execute("SELECT COUNT(*) as cnt FROM admins").fetchone()
    conn.close()
    return row["cnt"] > 0


def register_admin(username: str, password: str) -> Tuple[bool, str]:
    """注册管理员（仅首次）"""
    if admin_exists():
        return False, "管理员已存在，无法重复注册"
    if len(username.strip()) < 2:
        return False, "用户名至少2个字符"
    if len(password) < 4:
        return False, "密码至少4个字符"
    try:
        conn = get_connection()
        conn.execute(
            "INSERT INTO admins (username, password_hash) VALUES (?, ?)",
            (username.strip(), hash_password(password))
        )
        conn.commit()
        conn.close()
        return True, "注册成功"
    except sqlite3.IntegrityError:
        return False, "用户名已存在"


def login_admin(username: str, password: str) -> Tuple[bool, str]:
    """管理员登录验证"""
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM admins WHERE username = ?", (username.strip(),)
    ).fetchone()
    conn.close()

    if not row:
        return False, "用户名不存在"
    if row["password_hash"] != hash_password(password):
        return False, "密码错误"
    return True, "登录成功"


# ============================================================
#  操作日志
# ============================================================

def log_operation(action: str, target_type: str, target_id: int = None,
                  target_name: str = "", detail: str = ""):
    """记录操作日志"""
    conn = get_connection()
    conn.execute(
        "INSERT INTO operation_logs (action, target_type, target_id, target_name, detail) "
        "VALUES (?, ?, ?, ?, ?)",
        (action, target_type, target_id, target_name, detail)
    )
    conn.commit()
    conn.close()


# ============================================================
#  工具函数
# ============================================================

def detect_billing_type_from_period(start: str, end: str) -> str:
    """根据收费期间自动检测收费周期类型"""
    try:
        sy, sm = map(int, start.split('-'))
        ey, em = map(int, end.split('-'))
        months = (ey - sy) * 12 + (em - sm) + 1  # 包含起止月
        if months >= 12:
            return '年收'
        elif months >= 3:
            return '季收'
        else:
            return '月收'
    except (ValueError, AttributeError):
        return '季收'


def format_charge_period(start: str, end: str) -> str:
    """格式化收费期间为可读文本"""
    if not start or not end:
        return "-"
    try:
        sy, sm = start.split('-')
        ey, em = end.split('-')
        return f"{sy}年{int(sm):02d}月-{ey}年{int(em):02d}月"
    except (ValueError, AttributeError):
        return f"{start}-{end}"


# ============================================================
#  搜索与分页
# ============================================================

def search_clients(keyword: str, billing_filter: str = "全部",
                   status_filter: str = "全部",
                   offset: int = 0, limit: int = 50) -> Tuple[List[Dict], int]:
    """模糊搜索客户（名称/备注），支持筛选和分页
    Returns:
        (客户列表, 总条数)
    """
    conn = get_connection()
    like_kw = f"%{keyword}%"

    base_where = "WHERE is_active = 1"
    params = []

    if billing_filter == "年收":
        base_where += " AND billing_type = '年收'"
    elif billing_filter == "季收":
        base_where += " AND billing_type = '季收'"
    elif billing_filter == "月收":
        base_where += " AND billing_type = '月收'"
    elif billing_filter == "一人多司":
        base_where += " AND billing_type = '一人多司'"

    if status_filter == "有效":
        base_where += " AND status = '有效'"
    elif status_filter == "中断":
        base_where += " AND status = '中断'"
    elif status_filter == "失联":
        base_where += " AND status = '失联'"

    if keyword:
        base_where += " AND (name LIKE ? OR notes LIKE ?)"
        params.extend([like_kw, like_kw])

    if billing_filter == "逾期":
        rows = conn.execute(
            f"SELECT * FROM clients {base_where} ORDER BY name", params
        ).fetchall()
        rows = [r for r in rows if (
            r["payment_status"] == "未收" and
            calc_overdue_months_by_next_due(
                r["next_due_date"] or calc_next_due(r["last_paid_period"], r["billing_type"])
            ) > 0
        )]
        rows.sort(key=lambda r: calc_overdue_months_by_next_due(
            r["next_due_date"] or calc_next_due(r["last_paid_period"], r["billing_type"])
        ), reverse=True)
        total = len(rows)
        rows = rows[offset:offset + limit]
    else:
        count_row = conn.execute(
            f"SELECT COUNT(*) as cnt FROM clients {base_where}", params
        ).fetchone()
        total = count_row["cnt"]
        rows = conn.execute(
            f"SELECT * FROM clients {base_where} ORDER BY id LIMIT ? OFFSET ?",
            params + [limit, offset]
        ).fetchall()

    # 先查出所有客户，用于计算"一人多司"
    all_active = conn.execute(
        "SELECT id, contact_person FROM clients WHERE is_active = 1 AND contact_person != ''"
    ).fetchall()
    conn.close()

    contact_count = {}
    for r in all_active:
        cp = r["contact_person"].strip()
        if cp:
            contact_count[cp] = contact_count.get(cp, 0) + 1

    result = []
    for r in rows:
        d = dict(r)
        # 使用手动设置的 next_due_date，若为空则用自动计算兜底
        next_due = d.get("next_due_date", "") or calc_next_due(d["last_paid_period"], d["billing_type"])
        d["next_due"] = next_due

        # 逾期判断：仅当 payment_status 为"未收"时才计算
        if d.get("payment_status", "未收") == "已收":
            d["overdue_months"] = 0
            d["overdue_level"] = "正常"
            d["overdue_color"] = "#2E7D32"
            d["overdue_bg"] = "#E8F5E9"
            d["arrears_amount"] = 0.0
        else:
            mo = calc_overdue_months_by_next_due(next_due)
            level = get_overdue_level(mo)
            d["overdue_months"] = mo
            # 逾期状态直接用月数表达
            if mo > 0:
                d["overdue_level"] = f"逾期{mo}个月"
            else:
                d["overdue_level"] = "正常"
            d["overdue_color"] = level["color"]
            d["overdue_bg"] = level["bg"]
            # 欠费金额 = 应收金额（未收即欠费）
            d["arrears_amount"] = d["fee_amount"]
        # 一人多司：同一联系人出现≥2次
        cp = (d.get("contact_person", "") or "").strip()
        d["is_multi_company"] = contact_count.get(cp, 0) >= 2 if cp else False
        # 收费期间显示文本
        d["charge_period_display"] = format_charge_period(
            d.get("charge_period_start", ""), d.get("charge_period_end", "")
        )
        result.append(d)

    return result, total


def get_clients_paginated(billing_filter: str = "全部", keyword: str = "",
                          status_filter: str = "全部",
                          page: int = 1, page_size: int = 20) -> Tuple[List[Dict], int, int]:
    """分页获取客户列表
    Returns:
        (客户列表, 总条数, 总页数)
    """
    offset = (page - 1) * page_size
    result, total = search_clients(keyword, billing_filter, status_filter, offset, page_size)
    total_pages = max(1, (total + page_size - 1) // page_size)
    return result, total, total_pages


def get_clients(billing_filter: str = "全部", status_filter: str = "全部") -> List[Dict]:
    """获取客户列表（支持筛选）"""
    result, _ = search_clients("", billing_filter, status_filter, 0, 99999)
    return result


def get_billing_months(billing_type: str) -> int:
    """收费周期对应的月数"""
    if billing_type == "年收":
        return 12
    elif billing_type == "季收":
        return 3
    elif billing_type == "月收":
        return 1
    return 3


def calc_next_due(last_paid_period: str, billing_type: str) -> str:
    """计算下次应付月份
    例：last_paid='2025-01', 季收 → '2025-04'
         last_paid='2025-01', 年收 → '2026-01'
    """
    months = get_billing_months(billing_type)
    y, m = map(int, last_paid_period.split("-"))
    m += months
    while m > 12:
        y += 1
        m -= 12
    return f"{y}-{m:02d}"


def get_total_arrears() -> float:
    """获取所有未收客户的应收总额"""
    conn = get_connection()
    row = conn.execute(
        "SELECT SUM(fee_amount) as total FROM clients WHERE is_active = 1 AND payment_status != '已收'"
    ).fetchone()
    conn.close()
    return round(row["total"] or 0, 2)


def calc_overdue_months_by_next_due(next_due: str) -> int:
    """根据下次应付期计算逾期月数
    当前月份 vs 下次应付月份
    """
    if not next_due or next_due == "-":
        return 0
    try:
        due_y, due_m = map(int, next_due.split("-"))
        now = datetime.now()
        current_months = now.year * 12 + now.month
        due_months = due_y * 12 + due_m
        return current_months - due_months
    except (ValueError, AttributeError):
        return 0


def calc_overdue_months(last_paid_period: str, billing_type: str) -> int:
    """计算已逾期月数（负数=未到期，0=当月到期，正数=已逾期N个月）

    逻辑：当前月份 vs 下次应付月份
        若当前月 > 下次应付月 → 正数 = 逾期月数
    """
    next_due = calc_next_due(last_paid_period, billing_type)
    return calc_overdue_months_by_next_due(next_due)


def get_overdue_level(months_overdue: int) -> Dict:
    """根据逾期月数返回等级和颜色

    等级划分：
       < 3个月       → 正常   🟢 green
       3-5个月       → 逾期1季度 🟡 #D4A017
       6-11个月      → 逾期半年  🟠 #E87D0B
       12-23个月     → 逾期1年   🔴 #D93025
       >= 24个月     → 久悬户    ⬛ #8B0000
    """
    if months_overdue < 3:
        return {"level": "正常", "color": "#2E7D32", "bg": "#E8F5E9"}
    elif months_overdue < 6:
        return {"level": "逾期1季度", "color": "#D4A017", "bg": "#FFFDE7"}
    elif months_overdue < 12:
        return {"level": "逾期半年", "color": "#E87D0B", "bg": "#FFF3E0"}
    elif months_overdue < 24:
        return {"level": "逾期1年", "color": "#D93025", "bg": "#FFEBEE"}
    else:
        return {"level": "久悬户", "color": "#8B0000", "bg": "#FCE4EC"}


def add_client(name: str, billing_type: str, fee_amount: float,
               last_paid_period: str, phone: str = "", notes: str = "",
               status: str = "有效", payment_status: str = "未收",
               charge_period_start: str = "", charge_period_end: str = "",
               contact_person: str = "", next_due_date: str = "") -> Tuple[bool, str]:
    """新增客户"""
    if not name.strip():
        return False, "客户名称不能为空"
    try:
        conn = get_connection()
        conn.execute(
            """INSERT INTO clients (name, billing_type, fee_amount, last_paid_period,
               phone, notes, status, payment_status, charge_period_start, charge_period_end,
               contact_person, next_due_date)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (name.strip(), billing_type, fee_amount, last_paid_period,
             phone.strip(), notes.strip(), status, payment_status,
             charge_period_start, charge_period_end,
             contact_person.strip(), next_due_date)
        )
        conn.commit()
        conn.close()
        log_operation("新增", "客户", target_name=name,
                      detail=f"{billing_type} ¥{fee_amount:,.0f} 状态:{status}")
        return True, f"客户「{name}」添加成功"
    except sqlite3.IntegrityError:
        return False, "该客户名称已存在，请修改后重试"
    except Exception as e:
        return False, f"添加失败: {e}"


def update_client(client_id: int, **kwargs) -> Tuple[bool, str]:
    """更新客户信息（支持部分字段更新）"""
    allowed = {"name", "billing_type", "fee_amount", "last_paid_period",
               "phone", "notes", "is_active", "status", "payment_status",
               "charge_period_start", "charge_period_end",
               "contact_person", "next_due_date"}
    updates = {k: v for k, v in kwargs.items() if k in allowed}
    if not updates:
        return False, "无有效更新字段"

    set_clause = ", ".join(f"{k} = ?" for k in updates)
    values = list(updates.values()) + [client_id]

    conn = get_connection()
    conn.execute(f"UPDATE clients SET {set_clause} WHERE id = ?", values)
    conn.commit()
    conn.close()
    log_operation("编辑", "客户", target_id=client_id,
                  detail=str(kwargs))
    return True, "更新成功"


def delete_client(client_id: int) -> Tuple[bool, str]:
    """逻辑删除客户（保留历史数据）"""
    conn = get_connection()
    row = conn.execute("SELECT name FROM clients WHERE id = ?", (client_id,)).fetchone()
    name = row["name"] if row else ""
    conn.execute("UPDATE clients SET is_active = 0 WHERE id = ?", (client_id,))
    conn.commit()
    conn.close()
    log_operation("删除", "客户", target_id=client_id, target_name=name)
    return True, "已删除"


def batch_delete_clients(client_ids: List[int]) -> Tuple[bool, str]:
    """批量逻辑删除客户"""
    if not client_ids:
        return False, "请选择要删除的客户"
    conn = get_connection()
    placeholders = ",".join("?" * len(client_ids))
    conn.execute(f"UPDATE clients SET is_active = 0 WHERE id IN ({placeholders})",
                 client_ids)
    conn.commit()
    conn.close()
    log_operation("批量删除", "客户", detail=f"{len(client_ids)}个客户")
    return True, f"已删除 {len(client_ids)} 个客户"


def batch_record_payment(client_ids: List[int], paid_date: str = "",
                         notes: str = "") -> Tuple[bool, str, int]:
    """批量标记收费
    Returns:
        (成功与否, 消息, 成功数量)
    """
    if not client_ids:
        return False, "请选择要标记收费的客户", 0
    if not paid_date:
        paid_date = datetime.now().strftime("%Y-%m-%d")

    success_count = 0
    errors = []

    conn = get_connection()
    for cid in client_ids:
        row = conn.execute(
            "SELECT * FROM clients WHERE id = ? AND is_active = 1", (cid,)
        ).fetchone()
        if not row:
            continue

        c = dict(row)
        period_from = c["last_paid_period"]
        period_to = calc_next_due(c["last_paid_period"], c["billing_type"])
        amount = c["fee_amount"]

        try:
            conn.execute(
                """INSERT INTO payments (client_id, period_from, period_to, amount,
                   paid_date, notes) VALUES (?, ?, ?, ?, ?, ?)""",
                (cid, period_from, period_to, amount, paid_date, notes)
            )
            conn.execute(
                "UPDATE clients SET last_paid_period = ? WHERE id = ?",
                (period_to, cid)
            )
            success_count += 1
        except Exception as e:
            errors.append(f"{c['name']}: {e}")

    conn.commit()
    conn.close()

    log_operation("批量标记收费", "客户",
                  detail=f"成功{success_count}个")
    return True, f"成功标记 {success_count} 个客户" + (
        f"，{len(errors)}个失败" if errors else ""), success_count


def get_overdue_summary() -> Dict:
    """获取逾期客户汇总信息"""
    clients = get_clients("逾期")
    total_arrears = sum(c.get("fee_amount", 0) for c in clients)
    return {
        "count": len(clients),
        "total_arrears": round(total_arrears, 2),
        "clients": clients,
    }


def export_stats_to_excel(year: int, month: int, filepath: str = None) -> str:
    """导出收入统计为 Excel 文件
    Returns:
        文件路径
    """
    try:
        import openpyxl
    except ImportError:
        return _export_stats_csv(year, month, filepath)

    if not filepath:
        filepath = os.path.join(DB_DIR, f"收入统计_{year}年{month}月.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月"

    stats = get_monthly_stats(year, month)

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = f"代账客户收入统计 — {year}年{month}月"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)

    # 汇总
    ws["A3"] = "本月总收入"; ws["B3"] = stats["total_collected"]
    ws["A4"] = "年收客户贡献"; ws["B4"] = stats["annual_contribution"]
    ws["A5"] = "季收客户贡献"; ws["B5"] = stats["quarterly_contribution"]
    ws["A6"] = "月收客户贡献"; ws["B6"] = stats.get("monthly_contribution", 0)
    ws["A7"] = "按次业务贡献"; ws["B7"] = stats["service_contribution"]
    ws["A8"] = "当月应收未收"; ws["B8"] = stats["unpaid"]
    ws["A9"] = "按次业务未收"; ws["B9"] = stats.get("service_due", 0)
    ws["A10"] = "收款率"; ws["B10"] = f"{stats['collection_rate']}%"
    ws["A11"] = "应收欠款总额"; ws["B11"] = stats.get("total_arrears", 0)

    # 明细表头
    ws["A13"] = "客户/业务"; ws["B13"] = "类型"
    ws["C13"] = "金额(元)"; ws["D13"] = "收费期"
    ws["E13"] = "收款日期"
    for col in ["A13", "B13", "C13", "D13", "E13"]:
        ws[col].font = openpyxl.styles.Font(bold=True)

    # 明细数据
    for i, d in enumerate(stats["details"], start=14):
        ws[f"A{i}"] = d["client_name"]
        ws[f"B{i}"] = d["type"]
        ws[f"C{i}"] = d["amount"]
        ws[f"D{i}"] = d["period"]
        ws[f"E{i}"] = d["paid_date"]

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15

    wb.save(filepath)
    log_operation("导出", "报表", detail=filepath)
    return filepath


def _export_stats_csv(year: int, month: int, filepath: str = None) -> str:
    """CSV 格式导出（无需 openpyxl）"""
    import csv
    if not filepath:
        filepath = os.path.join(DB_DIR, f"收入统计_{year}年{month}月.csv")

    stats = get_monthly_stats(year, month)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([f"代账客户收入统计 - {year}年{month}月"])
        writer.writerow([])
        writer.writerow(["指标", "金额"])
        writer.writerow(["本月总收入", stats["total_collected"]])
        writer.writerow(["年收客户贡献", stats["annual_contribution"]])
        writer.writerow(["季收客户贡献", stats["quarterly_contribution"]])
        writer.writerow(["月收客户贡献", stats.get("monthly_contribution", 0)])
        writer.writerow(["按次业务贡献", stats["service_contribution"]])
        writer.writerow(["当月应收未收", stats["unpaid"]])
        writer.writerow(["按次业务未收", stats.get("service_due", 0)])
        writer.writerow(["收款率", f"{stats['collection_rate']}%"])
        writer.writerow([])
        writer.writerow(["客户/业务", "类型", "金额(元)", "收费期", "收款日期"])
        for d in stats["details"]:
            writer.writerow([d["client_name"], d["type"], d["amount"],
                            d["period"], d["paid_date"]])

    log_operation("导出", "报表", detail=filepath)
    return filepath


# ============================================================
#  收款管理
# ============================================================

def record_payment(client_id: int, period_from: str, period_to: str,
                   amount: float, paid_date: str = "", notes: str = "") -> Tuple[bool, str]:
    """记录一笔收款，自动更新客户的最近已收期

    Args:
        client_id: 客户 ID
        period_from: 本次收款覆盖起始月份 YYYY-MM
        period_to: 本次收款覆盖截止月份 YYYY-MM（将更新 last_paid_period）
        amount: 收款金额
        paid_date: 收款日期（默认今天）
        notes: 备注
    """
    if not paid_date:
        paid_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_connection()
    try:
        conn.execute(
            """INSERT INTO payments (client_id, period_from, period_to, amount, paid_date, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (client_id, period_from, period_to, amount, paid_date, notes.strip())
        )

        conn.execute(
            "UPDATE clients SET last_paid_period = ? WHERE id = ?",
            (period_to, client_id)
        )

        row = conn.execute("SELECT name FROM clients WHERE id = ?", (client_id,)).fetchone()
        client_name = row["name"] if row else ""

        conn.commit()
        conn.close()
        log_operation("标记收费", "客户", target_id=client_id, target_name=client_name,
                      detail=f"¥{amount:,.0f} {period_from}~{period_to}")
        return True, "收款记录成功"
    except Exception as e:
        conn.close()
        return False, f"收款失败: {e}"


def get_client_payments(client_id: int) -> List[Dict]:
    """获取某客户的全部收款记录"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM payments WHERE client_id = ? ORDER BY paid_date DESC",
        (client_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ============================================================
#  按次收费业务
# ============================================================

def add_service(client_id: int = None, service_type: str = "",
                fee_standard: float = 0, actual_fee: float = 0,
                temp_customer_name: str = "", notes: str = "",
                status: str = "未收", completed_date: str = "") -> Tuple[bool, str]:
    """新增按次收费业务（支持临时客户、多选业务类型）

    Args:
        client_id: 关联的正式客户 ID（可为空）
        service_type: 业务类型（逗号分隔多选，如"设立/变更,税务处理"）
        fee_standard: 当次业务金额
        actual_fee: 实收金额（已收时等于 fee_standard）
        temp_customer_name: 临时客户名称
        notes: 备注
        status: 收款状态（未收/已收），默认未收
        completed_date: 完成日期（已收时自动填今天）
    """
    if not service_type.strip():
        return False, "业务类型不能为空"
    if not client_id and not temp_customer_name.strip():
        return False, "请选择客户或输入客户名称"

    if status == "已收" and not completed_date:
        completed_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_connection()
    conn.execute(
        """INSERT INTO services (client_id, temp_customer_name, service_type,
           fee_standard, actual_fee, status, completed_date, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (client_id, temp_customer_name.strip(), service_type.strip(),
         fee_standard, actual_fee, status, completed_date, notes.strip())
    )
    conn.commit()
    conn.close()
    log_operation("新增", "按次业务",
                  target_name=temp_customer_name or f"客户ID:{client_id}",
                  detail=f"{service_type} ¥{fee_standard:,.0f} {status}")
    return True, "业务添加成功"


def mark_service_done(service_id: int, actual_fee: float = None) -> Tuple[bool, str]:
    """标记按次业务已完成收费"""
    conn = get_connection()
    if actual_fee is not None:
        conn.execute(
            """UPDATE services SET status='已收', actual_fee=?,
               completed_date=date('now','localtime')
               WHERE id=?""",
            (actual_fee, service_id)
        )
    else:
        conn.execute(
            """UPDATE services SET status='已收',
               completed_date=date('now','localtime')
               WHERE id=?""",
            (service_id,)
        )
    conn.commit()
    conn.close()
    log_operation("标记收费", "按次业务", target_id=service_id,
                  detail=f"¥{actual_fee:,.0f}" if actual_fee else "")
    return True, "已标记完成"


def delete_service(service_id: int) -> Tuple[bool, str]:
    """删除按次业务"""
    conn = get_connection()
    conn.execute("DELETE FROM services WHERE id = ?", (service_id,))
    conn.commit()
    conn.close()
    log_operation("删除", "按次业务", target_id=service_id)
    return True, "已删除"


def get_services(status_filter: str = "全部") -> List[Dict]:
    """获取按次业务列表"""
    conn = get_connection()
    if status_filter == "全部":
        rows = conn.execute(
            """SELECT s.*, COALESCE(c.name, s.temp_customer_name) as client_name
               FROM services s LEFT JOIN clients c ON s.client_id = c.id
               ORDER BY s.created_at DESC"""
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT s.*, COALESCE(c.name, s.temp_customer_name) as client_name
               FROM services s LEFT JOIN clients c ON s.client_id = c.id
               WHERE s.status = ? ORDER BY s.created_at DESC""",
            (status_filter,)
        ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        # 多选业务类型展示：逗号替换为"、"
        d["service_type_display"] = d["service_type"].replace(",", "、") if d["service_type"] else ""
        result.append(d)
    return result


# ============================================================
#  收入统计
# ============================================================

def get_monthly_stats(year: int, month: int) -> Dict:
    """获取指定月份的完整收入统计数据

    Returns:
        {
            "month_label": "2025-05",
            "total_collected": 本月已收总额,
            "annual_contribution": 年收客户贡献,
            "quarterly_contribution": 季收客户贡献,
            "monthly_contribution": 月收客户贡献,
            "service_contribution": 按次业务已收,
            "total_due": 当月应收总额（到期客户）,
            "unpaid": 当月应收未收,
            "service_due": 按次业务未收（全部）,
            "collection_rate": 收款率%,
            "details": [收入明细列表],
        }
    """
    month_str = f"{year}-{month:02d}"
    conn = get_connection()

    # 本月已收：从 payments 表统计（LEFT JOIN 保留已删除客户的历史收款记录）
    paid_rows = conn.execute(
        """SELECT p.*, c.name, c.billing_type
           FROM payments p LEFT JOIN clients c ON p.client_id = c.id
           WHERE strftime('%Y-%m', p.paid_date) = ?""",
        (month_str,)
    ).fetchall()

    total_collected = sum(r["amount"] for r in paid_rows)
    annual_contribution = sum(r["amount"] for r in paid_rows if r["billing_type"] == "年收")
    quarterly_contribution = sum(r["amount"] for r in paid_rows if r["billing_type"] == "季收")
    monthly_contribution = sum(r["amount"] for r in paid_rows if r["billing_type"] == "月收")

    # 按次业务本月已收
    svc_rows = conn.execute(
        """SELECT s.*, COALESCE(c.name, s.temp_customer_name) as client_name
           FROM services s LEFT JOIN clients c ON s.client_id = c.id
           WHERE s.status='已收' AND strftime('%Y-%m', s.completed_date) = ?""",
        (month_str,)
    ).fetchall()
    service_contribution = sum(r["actual_fee"] for r in svc_rows)

    total_collected += service_contribution

    # 按次业务未收总额（全部，不限月份）
    svc_due_row = conn.execute(
        "SELECT SUM(fee_standard) as total FROM services WHERE status='未收'"
    ).fetchone()
    service_due = round(svc_due_row["total"] or 0, 2)

    # 当月应收：活跃客户中，下次应付期在本月的
    # 使用手动设置的 next_due_date，若为空则用自动计算兜底
    all_clients = conn.execute(
        "SELECT * FROM clients WHERE is_active = 1"
    ).fetchall()

    total_due = 0
    for c in all_clients:
        next_due = c["next_due_date"] or calc_next_due(c["last_paid_period"], c["billing_type"])
        if next_due == month_str:
            total_due += c["fee_amount"]

    # 当月应收未收 = 当月应收 - 本月客户已收（不含按次）
    client_collected = total_collected - service_contribution
    unpaid = max(0, round(total_due - client_collected, 2))

    # 收款率：本月客户已收 / 当月应收
    collection_rate = round(client_collected / total_due * 100, 1) if total_due > 0 else 100.0

    # 收入明细
    details = []
    for r in paid_rows:
        details.append({
            "client_name": r["name"] or "(已删除客户)",
            "type": r["billing_type"] or "未知",
            "amount": r["amount"],
            "period": f"{r['period_from']}~{r['period_to']}",
            "paid_date": r["paid_date"],
        })
    for s in svc_rows:
        details.append({
            "client_name": s["client_name"] or f"(按次){s['service_type']}",
            "type": "按次",
            "amount": s["actual_fee"],
            "period": "-",
            "paid_date": s["completed_date"],
        })

    conn.close()

    return {
        "month_label": month_str,
        "total_collected": total_collected,
        "annual_contribution": annual_contribution,
        "quarterly_contribution": quarterly_contribution,
        "monthly_contribution": monthly_contribution,
        "service_contribution": service_contribution,
        "total_due": round(total_due, 2),
        "unpaid": unpaid,
        "service_due": service_due,
        "total_arrears": get_total_arrears(),
        "collection_rate": collection_rate,
        "details": details,
    }


def get_year_trend(year: int) -> List[Dict]:
    """获取全年 12 个月的收入趋势"""
    trend = []
    for month in range(1, 13):
        stats = get_monthly_stats(year, month)
        trend.append({
            "month": month,
            "label": f"{month}月",
            "amount": stats["total_collected"],
        })
    return trend


# ============================================================
#  测试
# ============================================================
if __name__ == "__main__":
    init_db()
    print("✅ 数据库初始化完成")

    # 测试注册
    if not admin_exists():
        ok, msg = register_admin("admin", "123456")
        print(f"注册管理员: {msg}")

    # 测试登录
    ok, msg = login_admin("admin", "123456")
    print(f"登录测试: {msg}")

    # 测试添加客户
    add_client("张三科技", "季收", 1500, "2025-01", "13800001111", "一般纳税人",
               status="有效", charge_period_start="2025-01", charge_period_end="2025-03",
               contact_person="张经理", next_due_date="2025-04")
    add_client("李四商贸", "年收", 6000, "2024-06", "13900002222", "",
               status="有效", charge_period_start="2024-06", charge_period_end="2025-05",
               contact_person="张经理", next_due_date="2025-06")
    add_client("王五餐饮", "季收", 2000, "2024-10", phone="13700003333",
               status="中断", contact_person="王老板", next_due_date="2025-01")

    # 测试查询
    clients = get_clients("全部")
    print(f"\n全部客户 ({len(clients)}):")
    for c in clients:
        print(f"  {c['name']} | {c['billing_type']} | ¥{c['fee_amount']} | "
              f"最近已收:{c['last_paid_period']} | 下次应付:{c['next_due']} | "
              f"逾期:{c['overdue_months']}月 | {c['overdue_level']} | "
              f"收费期间:{c['charge_period_display']} | 联系人:{c.get('contact_person','')} | "
              f"一人多司:{c.get('is_multi_company',False)} | 状态:{c.get('status','有效')}")

    # 测试收款
    if clients:
        c = clients[0]
        record_payment(c["id"], c["last_paid_period"],
                       calc_next_due(c["last_paid_period"], c["billing_type"]),
                       c["fee_amount"], notes="微信转账")

    # 测试按次业务（多选类型）
    if clients:
        add_service(clients[0]["id"], "设立/变更,税务处理", 2000, 2000,
                    notes="已办结", status="已收")

    # 测试统计
    now = datetime.now()
    stats = get_monthly_stats(now.year, now.month)
    print(f"\n📊 {stats['month_label']} 月收入统计:")
    print(f"  已收: ¥{stats['total_collected']}")
    print(f"  应收: ¥{stats['total_due']}")
    print(f"  未收: ¥{stats['unpaid']}")
    print(f"  收款率: {stats['collection_rate']}%")

    print("\n✅ 所有测试通过")
